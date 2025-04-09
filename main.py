from fastapi import FastAPI, Response
from fastapi.responses import StreamingResponse, JSONResponse
from pytrends.request import TrendReq
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import pandas as pd
from io import BytesIO

pd.set_option('future.no_silent_downcasting', True)

app = FastAPI()

@app.get("/")
def root():
    return {"status": "OK", "message": "API d'analyse comportementale opérationnelle"}

@app.get("/generate")
def generate_report():
    try:
        pytrends = TrendReq(hl='fr-FR', tz=360)

        SECTORS = {
            "Beauté": ["coiffeur", "soin visage", "épilation", "vernis", "institut beauté"],
            "Restauration": ["restaurant", "livraison repas", "UberEats", "menu midi", "réservation resto"],
            "Voyage & mobilité": ["billets d'avion", "location voiture", "Airbnb", "réservation hôtel", "train"],
            "Retail / Luxe": ["acheter chaussures", "sac à main", "montre luxe", "boutique mode", "soldes"],
            "Technologie & abonnements": ["abonnement Netflix", "désabonnement", "Spotify", "Disney+", "Prime Video"]
        }

        wb = Workbook()
        ws = wb.active
        ws.title = "Analyse Comportementale"

        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        positive_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        negative_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
        neutral_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

        headers = ["Secteur", "Mot-clé", "Score moyen", "Variation (%)", "Interprétation"]
        ws.append(headers)
        for col in range(1, 6):
            ws.cell(row=1, column=col).font = Font(bold=True)
            ws.cell(row=1, column=col).fill = header_fill
            ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")

        row_index = 2
        sentiments = []
        sector_summaries = {}

        for sector, keywords in SECTORS.items():
            pytrends.build_payload(keywords, cat=0, timeframe='now 7-d', geo='FR', gprop='')
            data = pytrends.interest_over_time()
            if 'isPartial' in data.columns:
                data = data.drop(columns='isPartial')

            deltas = []
            for keyword in keywords:
                if keyword in data.columns:
                    trend = data[keyword]
                    avg = round(trend.mean(), 1)
                    pct_change = round(((trend.iloc[-1] - trend.iloc[0]) / max(trend.iloc[0], 1)) * 100, 1)
                    deltas.append(pct_change)
                    impact = "↗️ Hausse" if pct_change > 10 else "↘️ Baisse" if pct_change < -10 else "➖ Stable"

                    ws.append([sector, keyword, avg, pct_change, impact])
                    color = positive_fill if pct_change > 10 else negative_fill if pct_change < -10 else neutral_fill
                    for col in range(1, 6):
                        ws.cell(row=row_index, column=col).fill = color
                        ws.cell(row=row_index, column=col).alignment = Alignment(horizontal="center")
                    row_index += 1

            sentiment = sum(deltas) / len(deltas) if deltas else 0
            sentiments.append(sentiment)
            summary_text = f"Résumé {sector} : "
            if sentiment > 10:
                summary_text += f"Tendance positive (variation moyenne : +{round(sentiment,1)}%)"
            elif sentiment < -10:
                summary_text += f"Tendance négative (variation moyenne : {round(sentiment,1)}%)"
            else:
                summary_text += f"Tendance stable (variation moyenne : {round(sentiment,1)}%)"
            sector_summaries[sector] = summary_text

        row_index += 2
        ws.append(["📌 Résumés sectoriels :"])
        ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=5)
        row_index += 1

        for summary in sector_summaries.values():
            ws.append([summary])
            ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=5)
            row_index += 1

        global_sentiment = sum(sentiments) / len(sentiments)
        global_summary = "🧠 Sentiment global : "
        if global_sentiment > 10:
            global_summary += f"positif (+{round(global_sentiment,1)}%)"
        elif global_sentiment < -10:
            global_summary += f"négatif ({round(global_sentiment,1)}%)"
        else:
            global_summary += f"stable ({round(global_sentiment,1)}%)"
        ws.append([global_summary])
        ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=5)

        # Save the workbook to a BytesIO stream
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Analyse_Secteurs_Comportement_{datetime.today().strftime('%Y%m%d')}.xlsx"
        headers = {
            'Content-Disposition': f'attachment; filename="{filename}"'
        }

        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})
